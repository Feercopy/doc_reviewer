from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any


MAX_SHEETS = 12
MAX_ROWS_PER_SHEET = 180
MAX_COLUMNS_PER_ROW = 60
MAX_CELL_TEXT_LENGTH = 2_000


def extract_workbook_snapshot(path: Path | str) -> dict[str, Any]:
    """Return a deterministic, bounded, JSON-serializable workbook snapshot."""
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl_unavailable") from exc

    workbook_path = Path(path).expanduser().resolve()
    formulas_workbook = openpyxl.load_workbook(workbook_path, data_only=False, read_only=True)
    values_workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        sheet_names = list(formulas_workbook.sheetnames)
        sheets = []
        for sheet_name in sheet_names[:MAX_SHEETS]:
            formula_sheet = formulas_workbook[sheet_name]
            values_sheet = values_workbook[sheet_name] if sheet_name in values_workbook.sheetnames else None
            sheets.append(_extract_sheet(formula_sheet=formula_sheet, values_sheet=values_sheet))

        return {
            "format": "xlsx_bounded_snapshot_v1",
            "source_filename": workbook_path.name,
            "limits": {
                "max_sheets": MAX_SHEETS,
                "max_rows_per_sheet": MAX_ROWS_PER_SHEET,
                "max_columns_per_row": MAX_COLUMNS_PER_ROW,
                "max_cell_text_length": MAX_CELL_TEXT_LENGTH,
            },
            "sheet_count": len(sheet_names),
            "sheets_truncated": len(sheet_names) > MAX_SHEETS,
            "sheets": sheets,
        }
    finally:
        formulas_workbook.close()
        values_workbook.close()


def _extract_sheet(*, formula_sheet: Any, values_sheet: Any | None) -> dict[str, Any]:
    max_row = formula_sheet.max_row or 0
    max_column = formula_sheet.max_column or 0
    rows = []
    for row_index in range(1, min(max_row, MAX_ROWS_PER_SHEET) + 1):
        row_cells = []
        for column_index in range(1, min(max_column, MAX_COLUMNS_PER_ROW) + 1):
            formula_cell = formula_sheet.cell(row=row_index, column=column_index)
            values_cell = values_sheet.cell(row=row_index, column=column_index) if values_sheet is not None else None
            record = _cell_record(
                address=getattr(formula_cell, "coordinate", _cell_address(row_index, column_index)),
                column_index=column_index,
                formula_value=getattr(formula_cell, "value", None),
                data_only_value=getattr(values_cell, "value", None) if values_cell is not None else None,
            )
            if record is not None:
                row_cells.append(record)
        if row_cells:
            rows.append({"row_number": row_index, "cells": row_cells})

    return {
        "name": formula_sheet.title,
        "dimensions": {
            "max_row": max_row,
            "max_column": max_column,
        },
        "rows_truncated": max_row > MAX_ROWS_PER_SHEET,
        "columns_truncated": max_column > MAX_COLUMNS_PER_ROW,
        "rows": rows,
    }


def _cell_record(
    *,
    address: str,
    column_index: int,
    formula_value: Any,
    data_only_value: Any,
) -> dict[str, Any] | None:
    if formula_value is None and data_only_value is None:
        return None

    record: dict[str, Any] = {
        "address": address,
        "column": column_index,
    }
    if isinstance(formula_value, str) and formula_value.startswith("="):
        record["formula"] = _serialize_value(formula_value)
        record["data_only_value"] = _serialize_value(data_only_value)
    else:
        record["value"] = _serialize_value(formula_value)
        if data_only_value != formula_value:
            record["data_only_value"] = _serialize_value(data_only_value)
    return record


def _serialize_value(value: Any) -> Any:
    if isinstance(value, (str, int, float, bool)) or value is None:
        return _redact_long_text(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return _redact_long_text(str(value))


def _redact_long_text(value: Any) -> Any:
    if isinstance(value, str) and len(value) > MAX_CELL_TEXT_LENGTH:
        return {
            "redacted": True,
            "reason": "cell_text_too_long",
            "length": len(value),
        }
    return value


def _cell_address(row_index: int, column_index: int) -> str:
    letters = ""
    column = column_index
    while column:
        column, remainder = divmod(column - 1, 26)
        letters = f"{chr(65 + remainder)}{letters}"
    return f"{letters}{row_index}"
